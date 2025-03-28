import os
from datetime import datetime
from .excel_core import ExcelCore
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from .pdf_service import PDFService 
from ..models.diploma_repository import DiplomaRepository
from dotenv import load_dotenv

load_dotenv()

class BaseExcelGenerator:
    def __init__(self, template_path: str = None, logo_path: str = None):
        self.logo_path = logo_path
        self.template_path = self._find_template(template_path)
        
    def _find_template(self, template_path: str, default_name: str = None) -> str:
        if template_path and os.path.exists(template_path):
            return template_path
            
        if default_name:
            script_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            default_template = os.path.join(script_dir, default_name)
            if os.path.exists(default_template):
                return default_template
                
        return None
    
    def _add_logo(self, excel: ExcelCore, cell_address: str):
        if not self.logo_path or not os.path.exists(self.logo_path):
            return
            
        worksheet = excel.sheet
        img = ExcelImage(self.logo_path)
        img.width = 150
        img.height = 90
        worksheet.add_image(img, cell_address)
        
        for row in range(1, 10):
            for col in range(1, 10):
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    
    def _adjust_column_widths(self, excel: ExcelCore, max_column_width: int = 40):
        for column in excel.sheet.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)

            for cell in column:
                if cell.value:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

            adjusted_width = min(max_column_width, (max_length + 2) * 1.2)
            excel.sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _adjust_row_heights(self, excel: ExcelCore, header_rows: tuple = None, min_height: int = 20):
        for row in excel.sheet.rows:
            if header_rows and header_rows[0] <= row[0].row <= header_rows[1]:
                continue
                
            max_height = 0
            for cell in row:
                if cell.value:
                    cell_text = str(cell.value)
                    col_letter = openpyxl.utils.get_column_letter(cell.column)
                    col_width = excel.sheet.column_dimensions[col_letter].width

                    chars_per_line = int(col_width / 1.2)
                    if chars_per_line <= 0:
                        chars_per_line = 1
                    
                    lines = 0
                    for line in cell_text.split('\n'):
                        lines += max(1, len(line) // chars_per_line + (1 if len(line) % chars_per_line > 0 else 0))
                    
                    row_height = lines * 15  # 15 пикселей на строку
                    if row_height > max_height:
                        max_height = row_height
            
            if max_height > 0:
                excel.sheet.row_dimensions[row[0].row].height = max(min_height, max_height)

    def _check_diploma_eligibility(self, assignments_results: list) -> tuple:
        if not assignments_results:
            return False, False
        
        all_passed = all(result.get("score", 0) >= 3 for result in assignments_results)
        with_honors = all(result.get("score", 0) == 5 for result in assignments_results)
        
        return all_passed, with_honors


class DiplomaGenerator(BaseExcelGenerator):  
    def __init__(self, template_path: str = None, logo_path: str = None):
        super().__init__(template_path, logo_path)
        self.template_path = self._find_template(template_path, "diploma.xlsx")
        # if not self.template_path and os.getenv("DEBUG") == "True":
        #     print("Предупреждение: Шаблон диплома не найден. Будет создан простой диплом.")
        # else:
        #     raise FileNotFoundError(f"Не загружен шаблон диплома: {self.template_path}")
    
    def _add_logo(self, excel: ExcelCore):
        super()._add_logo(excel, 'D1')
    
    def generate_diploma(self, student_data: dict, topic_name: str, 
                         assignments_results: list, output_path: str = None) -> str:
        eligible, with_honors = self._check_diploma_eligibility(assignments_results)
        if not eligible:
            raise ValueError("Студент не имеет права на получение диплома. "
                             "Все задания должны быть выполнены с оценкой не ниже 3.")
        
        if self.template_path and os.path.exists(self.template_path):
            excel = ExcelCore(self.template_path)
        else:
            excel = ExcelCore()
            excel.create_sheet("Диплом")
        
        self._add_logo(excel)

        student_name = student_data.get('full_name', '')
        excel.set_value("E11", student_name)
        excel.set_value("E13", topic_name)
        
        avg_score = sum(result.get("score", 0) for result in assignments_results) / len(assignments_results)
        excel.set_value("G15", f"{avg_score:.1f}")
        
        if with_honors:
            pass
        
        if not output_path:
            student_name_file = student_data.get('full_name', '').replace(' ', '_')
            output_path = f"diploma_{student_name_file}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

        for row in excel.sheet.rows:
            for cell in row:
                if 6 <= cell.row <= 8 and 3 <= cell.column <= 7:
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                elif (cell.row == 11 and cell.column == 5) or (cell.row == 13 and cell.column == 5) or (cell.row == 15 and cell.column == 7):
                    cell.font = openpyxl.styles.Font(size=12)
                    cell.alignment = openpyxl.styles.Alignment(horizontal='right', vertical='center', wrap_text=True)
                else:
                    cell.font = openpyxl.styles.Font(size=12)
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        self._adjust_column_widths(excel)
        self._adjust_row_heights(excel, (6, 8))
        
        excel.save(output_path)
        return output_path


class DiplomaAppendixGenerator(BaseExcelGenerator):    
    def __init__(self, template_path: str = None, logo_path: str = None):
        super().__init__(template_path, logo_path)
        self.template_path = self._find_template(template_path, "diploma-addition.xlsx")
            
    def _add_logo(self, excel: ExcelCore):
        super()._add_logo(excel, 'C1')
    
    def generate_appendix(self, student_data: dict, topic_name: str, 
                          assignments_results: list, output_path: str = None,
                          issued_by: str = "Выдано") -> str:
        if self.template_path and os.path.exists(self.template_path):
            excel = ExcelCore(self.template_path)
        else:
            excel = ExcelCore()
            excel.create_sheet("Приложение к диплому")

        self._add_logo(excel)
        excel.set_value("B10", f"Результаты по теме: {topic_name}")
        excel.set_value("B13", "№")
        excel.set_value("C13", "Название задания")
        excel.set_value("D13", "Оценка")
        excel.set_value("E13", "Время (мин)")
        
        total_score = 0
        total_time = 0
        
        for idx, result in enumerate(assignments_results, start=1):
            row = 13 + idx
            excel.set_value(f"B{row}", idx)
            excel.set_value(f"C{row}", result.get("name", ""))
            excel.set_value(f"D{row}", result.get("score", 0))
            excel.set_value(f"E{row}", result.get("time_spent", 0))
            
            total_score += result.get("score", 0)
            total_time += result.get("time_spent", 0)
        
        summary_row = 14 + len(assignments_results)
        excel.set_value(f"B{summary_row}", "Итого:")
        excel.set_value(f"D{summary_row}", total_score)
        excel.set_value(f"E{summary_row}", total_time)
        
        avg_score = total_score / len(assignments_results) if assignments_results else 0
        excel.set_value(f"C{summary_row + 1}", f"Средний балл: {avg_score:.1f}")
        
        signature_row = summary_row + 3
        
        full_name = student_data.get('full_name', '')
        initials = ""
        if full_name:
            name_parts = full_name.split()
            if len(name_parts) >= 2:
                surname = name_parts[0]
                initials = "".join([part[0] + "." for part in name_parts[1:]])
                full_name_short = f"{surname} {initials}"
            else:
                full_name_short = full_name
        else:
            full_name_short = ""
        
        excel.set_value(f"B{signature_row}", f"{issued_by}")
        excel.set_value(f"C{signature_row}", full_name_short)
        underline = "_" * 6
        excel.set_value(f"D{signature_row}", underline)
        
        current_date = datetime.now().strftime("%d.%m.%Y")
        excel.set_value(f"E{signature_row}", current_date)
        
        if not output_path:
            student_name_file = student_data.get('full_name', '').replace(' ', '_')
            output_path = f"diploma_appendix_{student_name_file}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

        for row in excel.sheet.rows:
            for cell in row:
                if 6 <= cell.row <= 8 and 2 <= cell.column <= 6:
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                else:
                    cell.font = openpyxl.styles.Font(size=12)
                    if row[0].row >= 13:
                        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

        self._adjust_column_widths(excel, 20)
        excel.sheet.column_dimensions['C'].width = 20 
        
        self._adjust_row_heights(excel, (6, 8), 25)
        
        for row_idx in range(14, 14 + len(assignments_results)):
            excel.sheet.row_dimensions[row_idx].height = 75
        
        excel.save(output_path)
        return output_path


class DiplomaService:   
    def __init__(self, logo_path: str = None, 
                 diploma_template: str = None, 
                 appendix_template: str = None):
        self.diploma_generator = DiplomaGenerator(
            template_path=diploma_template,
            logo_path=logo_path
        )
        
        self.appendix_generator = DiplomaAppendixGenerator(
            template_path=appendix_template,
            logo_path=logo_path
        )
    
    def _check_diploma_eligibility(self, assignments_results: list) -> tuple:
        return self.diploma_generator._check_diploma_eligibility(assignments_results)
    
    def export_to_pdf(self, excel_path: str, output_pdf_path: str = None) -> str:
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Файл Excel не найден: {excel_path}")
        
        if not output_pdf_path:
            output_pdf_path = os.path.splitext(excel_path)[0] + ".pdf"
        
        pdf_service = PDFService(timeout=60)
        
        pdf_path = pdf_service.convert_to_pdf(excel_path)
        
        if pdf_path and os.path.exists(pdf_path):
            if pdf_path != output_pdf_path and os.path.exists(pdf_path):
                try:
                    os.rename(pdf_path, output_pdf_path)
                    return output_pdf_path
                except:
                    return pdf_path
            return pdf_path
        return excel_path
    
    def generate_diploma_with_appendix(self, student_data: dict, topic_name: str, 
                                      assignments_results: list, output_dir: str = None,
                                      issued_by: str = "Выдано") -> dict:
        eligible, with_honors = self._check_diploma_eligibility(assignments_results)
        if not eligible:
            raise ValueError("Студент не имеет права на получение диплома. "
                             "Все задания должны быть выполнены с оценкой не ниже 3.")
        
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        student_name_file = student_data.get('full_name', '').replace(' ', '_')
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        
        if output_dir:
            diploma_excel_path = os.path.join(output_dir, f"diploma_{student_name_file}_{timestamp}.xlsx")
            diploma_pdf_path = os.path.join(output_dir, f"diploma_{student_name_file}_{timestamp}.pdf")
            appendix_excel_path = os.path.join(output_dir, f"diploma_appendix_{student_name_file}_{timestamp}.xlsx")
            appendix_pdf_path = os.path.join(output_dir, f"diploma_appendix_{student_name_file}_{timestamp}.pdf")
        else:
            diploma_excel_path = f"diploma_{student_name_file}_{timestamp}.xlsx"
            diploma_pdf_path = f"diploma_{student_name_file}_{timestamp}.pdf"
            appendix_excel_path = f"diploma_appendix_{student_name_file}_{timestamp}.xlsx"
            appendix_pdf_path = f"diploma_appendix_{student_name_file}_{timestamp}.pdf"
        
        self.diploma_generator.generate_diploma(
            student_data=student_data,
            topic_name=topic_name,
            assignments_results=assignments_results,
            output_path=diploma_excel_path
        )
        self.appendix_generator.generate_appendix(
            student_data=student_data,
            topic_name=topic_name,
            assignments_results=assignments_results,
            output_path=appendix_excel_path,
            issued_by=issued_by
        )
       
        diploma_pdf = None
        appendix_pdf = None
        
        try:
            diploma_pdf = self.export_to_pdf(diploma_excel_path, diploma_pdf_path)
        except Exception as e:
            print(f"Ошибка при экспорте диплома в PDF: {str(e)}")
        
        try:
            appendix_pdf = self.export_to_pdf(appendix_excel_path, appendix_pdf_path)
        except Exception as e:
            print(f"Ошибка при экспорте приложения в PDF: {str(e)}")
        
        return {
            "diploma_excel": diploma_excel_path,
            "appendix_excel": appendix_excel_path,
            "diploma_pdf": diploma_pdf,
            "appendix_pdf": appendix_pdf
        }
    
    def generate_diploma_by_diploma_id(self, diploma_id: int, output_dir: str = None, 
                                      issued_by: str = "Выдано") -> dict:
        diploma_data = DiplomaRepository.get_diploma_by_id(diploma_id)
        if not diploma_data:
            raise ValueError(f"Диплом с ID {diploma_id} не найден в базе данных")
        
        student_data = {
            "full_name": diploma_data["student"]["full_name"],
            "email": diploma_data["student"]["email"]
        }
        
        topic_name = diploma_data["topic_title"]
        assignments_results = []
        
        if isinstance(diploma_data.get("tasks"), dict):
            for task_name, score in diploma_data["tasks"].items():
                if isinstance(score, (int, float)):
                    assignments_results.append({
                        "name": task_name,
                        "score": score,
                        "time_spent": 0  # Здесь можно добавить логику для времени, если нужно
                    })
        else:
            print(f"Предупреждение: для диплома {diploma_id} поле 'tasks' не является словарем.")
            raise AttributeError("Нет информации о результатах для диплома")

        return self.generate_diploma_with_appendix(
            student_data=student_data,
            topic_name=topic_name,
            assignments_results=assignments_results,
            output_dir=output_dir,
            issued_by=issued_by
        )    
    
    def generate_diploma_by_user_and_topic(self, user_id: int, topic_id: int, 
                                          output_dir: str = None, 
                                          issued_by: str = "Выдано") -> dict:
        user_data = DiplomaRepository.get_user_by_id(user_id)
        if not user_data:
            raise ValueError(f"Пользователь с ID {user_id} не найден в базе данных")

        topic_data = DiplomaRepository.get_topic_by_id(topic_id)
        if not topic_data:
            raise ValueError(f"Тема с ID {topic_id} не найдена в базе данных")
        
        student_data = {
            "full_name": user_data["full_name"],
            "email": user_data["email"]
        }
        
        topic_name = topic_data["title"]
        assignments_results = []
        performed_tasks = DiplomaRepository.get_performed_tasks_by_user_id(user_id)
        if performed_tasks:
            for task in performed_tasks:
                if task.get("grade") is not None:
                    task_data = {
                        "name": f"Задание {task.get('task_id', 'б/н')}", 
                        "score": task.get("grade", 0),
                        "time_spent": 0
                    }
                    assignments_results.append(task_data)
        
        if not assignments_results:
            if os.getenv("DEBUG") == "True":
                print(f"Предупреждение: для пользователя {user_id} не найдены выполненные задания. Используем тестовые данные.")
                assignments_results = [
                    {"name": "Тестовое задание 1", "score": 5, "time_spent": 60},
                    {"name": "Тестовое задание 2", "score": 4, "time_spent": 90}
                ]
            else:
                raise AttributeError("Нет информации о результатах для диплома")

        return self.generate_diploma_with_appendix(
            student_data=student_data,
            topic_name=topic_name,
            assignments_results=assignments_results,
            output_dir=output_dir,
            issued_by=issued_by
        )    
    