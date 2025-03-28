import openpyxl
from openpyxl.workbook import Workbook

class ExcelCore:
    def __init__(self, file_path: str = None):
        self.file_path = file_path
        if file_path:
            self.workbook = openpyxl.load_workbook(file_path)
        else:
            self.workbook = Workbook()
        self.sheet = self.workbook.active

    def save(self, file_path: str = None):
        path = file_path if file_path else self.file_path
        if not path:
            raise ValueError("Не указан путь для сохранения файла")
        self.workbook.save(path)

    def set_value(self, cell: str, value):
        self.sheet[cell] = value

    def get_value(self, cell: str):
        return self.sheet[cell].value

    def select_sheet(self, sheet_name: str):
        if sheet_name in self.workbook.sheetnames:
            self.sheet = self.workbook[sheet_name]
        else:
            raise ValueError(f"Лист '{sheet_name}' не найден")

    def create_sheet(self, sheet_name: str):
        self.sheet = self.workbook.create_sheet(title=sheet_name)

    def list_sheets(self):
        return self.workbook.sheetnames
    